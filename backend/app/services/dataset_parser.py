"""Pure parsing helpers for dataset import (no DB access)."""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from app.core.exceptions import ValidationError

_EXPECTED_KEYS = ("expected", "answer", "label", "output", "target", "ground_truth")
_SUPPORTED = ("csv", "tsv", "json", "jsonl", "xlsx")
_CHAT_ROLES = ("system", "user", "assistant")

_TYPE_ALIASES = {
    "string": "string",
    "text": "string",
    "str": "string",
    "number": "number",
    "numeric": "number",
    "float": "number",
    "integer": "integer",
    "int": "integer",
    "boolean": "boolean",
    "bool": "boolean",
    "json": "json",
    "array": "array",
    "list": "array",
    "object": "object",
    "dict": "object",
}


def _decode_bytes(raw_bytes: bytes) -> str:
    """Decode raw text bytes: UTF-8 (BOM-safe), then GBK, then UTF-16 (BOM-only)."""
    for encoding in ("utf-8-sig", "gbk"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    if raw_bytes.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            return raw_bytes.decode("utf-16")
        except UnicodeDecodeError:
            pass
    raise ValidationError(
        "Cannot decode file: expected UTF-8, GBK/GB2312 or UTF-16 encoding"
    )


def infer_format(filename: str | None, fmt: str | None) -> str:
    """Resolve the upload format: explicit value wins, else extension inference."""
    fmt = (fmt or "").strip().lower()
    if fmt:
        return fmt
    if filename:
        ext = filename.rsplit(".", 1)[-1].lower()
        fmt = {"jsonl": "jsonl", "json": "json", "csv": "csv", "tsv": "tsv", "xlsx": "xlsx"}.get(ext, "")
    return fmt or "json"


def _validate_magic(raw_bytes: bytes, fmt: str) -> None:
    """Reject files whose declared format does not match their content signature."""
    if fmt == "xlsx":
        if not raw_bytes.startswith(b"PK"):
            raise ValidationError("File is not a valid XLSX (missing zip signature)")
        return
    if fmt in ("json", "jsonl"):
        text = _decode_bytes(raw_bytes)
        stripped = text.lstrip("\ufeff \t\r\n")
        if not stripped:
            return
        if fmt == "json" and stripped[0] not in "[{":
            raise ValidationError("JSON must start with '[' or '{'")
        if fmt == "jsonl" and stripped[0] != "{":
            raise ValidationError("JSONL must start with '{' on the first line")


def parse_dataset(raw_bytes: bytes, fmt: str) -> list[dict]:
    fmt = (fmt or "").strip().lower()
    if fmt not in _SUPPORTED:
        raise ValidationError(f"Unsupported format: {fmt!r}")
    _validate_magic(raw_bytes, fmt)

    if fmt in ("csv", "tsv"):
        text = _decode_bytes(raw_bytes)
        delimiter = "\t" if fmt == "tsv" else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValidationError(f"{fmt.upper()} has no header row")
        rows = [dict(r) for r in reader]
    elif fmt == "json":
        text = _decode_bytes(raw_bytes)
        try:
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValidationError(f"Invalid JSON: {exc}") from exc
        if isinstance(data, dict):
            for key in ("data", "rows"):
                if isinstance(data.get(key), list):
                    data = data[key]
                    break
        if not isinstance(data, list):
            raise ValidationError("JSON must be a list of objects")
        rows = data
    elif fmt == "jsonl":
        text = _decode_bytes(raw_bytes)
        rows = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValidationError(f"Invalid JSONL line: {exc}") from exc
    else:  # xlsx
        rows = _parse_xlsx(raw_bytes)

    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            raise ValidationError(f"Row {i} is not a JSON object")
    return rows


def _parse_xlsx(raw_bytes: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure means a bad file
        raise ValidationError(f"Invalid XLSX file: {exc}") from exc
    sheet = workbook.worksheets[0] if workbook.worksheets else None
    if sheet is None:
        raise ValidationError("XLSX has no sheets")
    values_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(values_iter)
    except StopIteration:
        headers = None
    if not headers:
        raise ValidationError("XLSX has no header row")
    rows: list[dict] = []
    for values in values_iter:
        row = {
            str(header): value
            for header, value in zip(headers, values)
            if header is not None
        }
        rows.append(row)
    return rows


def infer_schema(rows: list[dict]) -> list[str]:
    schema: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in schema:
                schema.append(key)
    return schema


def compute_stats(rows: list[dict]) -> dict:
    columns = infer_schema(rows)
    null_counts = {col: 0 for col in columns}
    for row in rows:
        for col in columns:
            if col not in row or _is_blank(row.get(col)):
                null_counts[col] += 1
    return {
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "null_counts": null_counts,
    }


def _field_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("["):
            try:
                value = json.loads(text)
            except json.JSONDecodeError as exc:
                raise ValidationError(f"Invalid field list JSON: {exc}") from exc
        else:
            return [part.strip() for part in text.split(",") if part.strip()]
    if isinstance(value, (list, tuple)):
        return [str(item).strip() for item in value if str(item).strip()]
    raise ValidationError("Field lists must be arrays or comma-separated strings")


def _json_object(value: Any, label: str) -> dict:
    if value is None:
        return {}
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValidationError(f"Invalid {label} JSON: {exc}") from exc
    if not isinstance(value, dict):
        raise ValidationError(f"{label} must be a JSON object")
    return value


def build_dataset_contract(
    rows: list[dict],
    *,
    task_type: str | None = None,
    input_fields: Any = None,
    expected_fields: Any = None,
    metadata_fields: Any = None,
    required_fields: Any = None,
    field_types: Any = None,
    answer_policy: Any = None,
    contract: Any = None,
    sensitive_fields: Any = None,
    structured_chat: Any = None,
) -> dict:
    """Build a lightweight dataset contract for import and validation."""
    payload = _json_object(contract, "contract")
    columns = infer_schema(rows)
    nested_mapping = payload.get("field_mapping")
    if not isinstance(nested_mapping, dict):
        nested_mapping = {}

    mapped_input = _field_list(
        input_fields
        if input_fields is not None
        else payload.get("input_fields", nested_mapping.get("input_fields"))
    )
    mapped_expected = _field_list(
        expected_fields
        if expected_fields is not None
        else payload.get("expected_fields", nested_mapping.get("expected_fields"))
    )
    mapped_metadata = _field_list(
        metadata_fields
        if metadata_fields is not None
        else payload.get("metadata_fields", nested_mapping.get("metadata_fields"))
    )
    mapped_sensitive = _field_list(
        sensitive_fields
        if sensitive_fields is not None
        else payload.get("sensitive_fields")
    )
    if structured_chat is not None:
        if isinstance(structured_chat, str):
            normalized_chat = structured_chat.strip().lower() in ("1", "true", "yes", "on")
        else:
            normalized_chat = bool(structured_chat)
    else:
        raw_chat = payload.get("structured_chat", False)
        normalized_chat = (
            raw_chat.strip().lower() in ("1", "true", "yes", "on")
            if isinstance(raw_chat, str)
            else bool(raw_chat)
        )

    if not mapped_expected:
        for key in _EXPECTED_KEYS:
            variants = [col for col in columns if col.lower() == key]
            if not variants:
                continue
            exact = [col for col in variants if col == key]
            mapped_expected.append(exact[0] if exact else variants[0])
    if not mapped_input:
        excluded = set(mapped_expected) | set(mapped_metadata)
        mapped_input = [col for col in columns if col not in excluded]

    roles = {
        "input": set(mapped_input),
        "expected": set(mapped_expected),
        "metadata": set(mapped_metadata),
    }
    overlap = (
        (roles["input"] & roles["expected"])
        | (roles["input"] & roles["metadata"])
        | (roles["expected"] & roles["metadata"])
    )
    if overlap:
        field = sorted(overlap)[0]
        raise ValidationError(f"Field mapped to multiple roles: {field}")

    for field in mapped_expected:
        if field not in columns and not any(_source_has_field(row, field) for row in rows):
            raise ValidationError(f"Expected field '{field}' is not present in source")

    try:
        schema_version = int(payload.get("schema_version", 1) or 1)
    except (TypeError, ValueError) as exc:
        raise ValidationError("schema_version must be an integer") from exc

    normalized = {
        "schema_version": schema_version,
        "task_type": task_type or payload.get("task_type") or "qa",
        "input_fields": mapped_input,
        "expected_fields": mapped_expected,
        "metadata_fields": mapped_metadata,
        "required_fields": _field_list(
            required_fields if required_fields is not None else payload.get("required_fields")
        ),
        "field_types": _json_object(
            field_types if field_types is not None else payload.get("field_types"),
            "field_types",
        ),
        "answer_policy": _json_object(
            answer_policy if answer_policy is not None else payload.get("answer_policy"),
            "answer_policy",
        ),
        "sensitive_fields": mapped_sensitive,
        "structured_chat": normalized_chat,
    }
    normalized["field_mapping"] = {
        "input_fields": normalized["input_fields"],
        "expected_fields": normalized["expected_fields"],
        "metadata_fields": normalized["metadata_fields"],
    }
    return normalized


def _source_has_field(row: dict, field: str) -> bool:
    if field in row and not _is_blank(row[field]):
        return True
    expected = row.get("expected")
    return isinstance(expected, dict) and not _is_blank(expected.get(field))


def _is_blank(value: Any) -> bool:
    """Empty or whitespace-only values count as blank."""
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def validate_required_fields(row: dict, contract: dict, row_idx: int) -> list[str]:
    issues: list[str] = []
    for field in contract.get("required_fields", []) or []:
        if not _source_has_field(row, field):
            issues.append(f"Row {row_idx} missing required field: {field}")
    return issues


def collect_required_field_errors(rows: list[dict], contract: dict) -> list[dict]:
    """Structured row-level required-field errors for import reports."""
    errors: list[dict] = []
    for i, row in enumerate(rows):
        for field in contract.get("required_fields", []) or []:
            if not _source_has_field(row, field):
                errors.append(
                    {
                        "row": i,
                        "field": field,
                        "message": f"Row {i} missing required field: {field}",
                    }
                )
    return errors


def _field_value(row: dict, field: str) -> Any:
    if field in row:
        return row[field]
    expected = row.get("expected")
    if isinstance(expected, dict):
        return expected.get(field)
    return None


def _value_matches_type(value: Any, type_name: str) -> bool:
    if value is None or _is_blank(value):
        return True  # blankness is handled by required-field checks
    if type_name == "string":
        return isinstance(value, str)
    if type_name == "number":
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            try:
                float(value)
                return True
            except ValueError:
                return False
        return False
    if type_name == "integer":
        if isinstance(value, bool):
            return False
        if isinstance(value, int):
            return True
        if isinstance(value, str):
            try:
                return float(value).is_integer()
            except ValueError:
                return False
        return False
    if type_name == "boolean":
        if isinstance(value, bool):
            return True
        if isinstance(value, str):
            return value.strip().lower() in ("true", "false", "1", "0")
        return False
    if type_name == "array":
        return isinstance(value, list)
    if type_name == "object":
        return isinstance(value, dict)
    if type_name == "json":
        return True
    return True


def collect_field_type_errors(rows: list[dict], contract: dict) -> list[dict]:
    """Validate declared ``field_types`` against every row's values."""
    errors: list[dict] = []
    field_types = contract.get("field_types", {}) or {}
    for field, declared in field_types.items():
        type_name = _TYPE_ALIASES.get(str(declared).strip().lower())
        if type_name is None:
            errors.append(
                {
                    "row": None,
                    "field": field,
                    "message": f"Unsupported field type '{declared}' for field '{field}'",
                }
            )
            continue
        for i, row in enumerate(rows):
            value = _field_value(row, field)
            if value is None:
                continue
            if not _value_matches_type(value, type_name):
                errors.append(
                    {
                        "row": i,
                        "field": field,
                        "message": f"Row {i}: field '{field}' is not {declared}",
                    }
                )
    return errors


def collect_chat_structure_errors(rows: list[dict], contract: dict) -> list[dict]:
    """Validate multi-turn messages / few-shot examples when structured chat is on."""
    if not contract.get("structured_chat"):
        return []
    errors: list[dict] = []
    for i, row in enumerate(rows):
        messages = row.get("messages")
        if messages is not None:
            if not isinstance(messages, list):
                errors.append(
                    {
                        "row": i,
                        "field": "messages",
                        "message": f"Row {i}: 'messages' must be a list",
                    }
                )
            else:
                for j, item in enumerate(messages):
                    valid = (
                        isinstance(item, dict)
                        and item.get("role") in _CHAT_ROLES
                        and isinstance(item.get("content"), str)
                    )
                    if not valid:
                        errors.append(
                            {
                                "row": i,
                                "field": "messages",
                                "message": (
                                    f"Row {i}: messages[{j}] must be "
                                    "{role: system|user|assistant, content: str}"
                                ),
                            }
                        )
        examples = row.get("examples")
        if examples is not None:
            if not isinstance(examples, list):
                errors.append(
                    {
                        "row": i,
                        "field": "examples",
                        "message": f"Row {i}: 'examples' must be a list",
                    }
                )
            else:
                for j, item in enumerate(examples):
                    if not isinstance(item, (str, dict)):
                        errors.append(
                            {
                                "row": i,
                                "field": "examples",
                                "message": (
                                    f"Row {i}: examples[{j}] must be a string or object"
                                ),
                            }
                        )
    return errors


def collect_import_errors(rows: list[dict], contract: dict) -> list[dict]:
    """All row-level import errors: required fields first, then type mismatches."""
    return (
        collect_required_field_errors(rows, contract)
        + collect_field_type_errors(rows, contract)
        + collect_chat_structure_errors(rows, contract)
    )


def split_input_expected(row: dict, contract: dict | None = None) -> tuple[dict, dict | None]:
    """Separate a dataset row into input fields and expected answer fields.

    Strategy:
    1. If the row contains an explicit ``expected`` key (a dict/list), treat it as
       the full expected structure and put everything else into input.
    2. Otherwise, scan for known answer-key names and collect them into a single
       expected dict. All other keys become input.
    3. If no expected key is found, treat the entire row as input with no expected.

    This handles common formats:

    * ``{"question": "...", "answer": "北京"}`` → input={question}, expected={answer:北京}
    * ``{"prompt": "...", "answer": "北京", "reasoning": "..."}`` → input={prompt}, expected={answer:北京, reasoning:...}
    * ``{"text": "...", "label": "正面"}`` → input={text}, expected={label:正面}
    * ``{"question": "...", "expected": {"answer": "北京"}}`` → input={question}, expected={answer:北京}
    """
    source = dict(row)
    mapping = (contract or {}).get("field_mapping", contract or {})
    input_fields = mapping.get("input_fields") or []
    expected_fields = mapping.get("expected_fields") or []
    metadata_fields = mapping.get("metadata_fields") or []

    if input_fields or expected_fields or metadata_fields:
        input_data = {field: source[field] for field in input_fields if field in source}
        metadata = {field: source[field] for field in metadata_fields if field in source}
        if metadata:
            input_data["_metadata"] = metadata

        expected: dict = {}
        nested_expected = source.get("expected")
        for field in expected_fields:
            if field == "expected" and "expected" in source:
                expected_val = source["expected"]
                if isinstance(expected_val, dict):
                    expected.update(expected_val)
                elif isinstance(expected_val, list):
                    expected["answer"] = expected_val
                else:
                    expected["answer"] = expected_val
            elif field in source:
                expected[field] = source[field]
            elif isinstance(nested_expected, dict) and field in nested_expected:
                expected[field] = nested_expected[field]
        return input_data, expected or None

    # Step 1: Check for explicit "expected" key (could be a dict or list).
    if "expected" in source:
        expected_val = source.pop("expected")
        if isinstance(expected_val, dict):
            return source, expected_val
        elif isinstance(expected_val, list):
            # Wrap list of answers into a single expected dict.
            return source, {"answer": expected_val}
        else:
            return source, {"answer": expected_val}

    # Step 2: Look for known answer keys and collect them.
    expected_keys = set()
    lower_to_original = {key.lower(): key for key in source}
    for key in source:
        if key.lower() in _EXPECTED_KEYS:
            expected_keys.add(lower_to_original[key.lower()])

    if expected_keys:
        expected = {}
        remaining = {}
        for key, value in source.items():
            if key in expected_keys:
                expected[key] = value
            else:
                remaining[key] = value
        return remaining, expected if expected else None

    # Step 3: No expected fields found — treat entire row as input.
    return source, None
